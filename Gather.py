
# This example requires the 'members' and 'message_content' privileged intents to function.

import discord
from discord import ButtonStyle, ui, app_commands,webhook
from datetime import datetime, timezone, timedelta
from discord.ext import tasks, commands
import random
import asyncio
import dice
import pandas as pd
import gspread
import typing
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials

blacklist = ["1076422611325702236","1093041645114622004","1032532645655105536","931562992041099275","1081170888063459368","872546782700273694","1066977490339373076","883332312484446229"]
description = '''テスト用botです'''

#スプレッドシート関連
scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
json = 'strl-380010-d9b3efdea4a1.json'
credentials = ServiceAccountCredentials.from_json_keyfile_name(json, scope)
gc = gspread.authorize(credentials)
SPREADSHEET_KEY = '1vXe0TvwhgoOypM4xGF-fkQPYSNaLwTJVN06LmonCfoA'
workbook = gc.open_by_key(SPREADSHEET_KEY)
datasheet = workbook.worksheet("Nユーザーデータ")
Fdatasheet = workbook.worksheet("Fユーザーデータ")
COMdatasheet = workbook.worksheet("コマンド登録")

target = ""

th_name = {}
food_count = []

JST = timezone(timedelta(hours=+9), 'JST')
now = datetime.now(JST)
dad = now.date().strftime('%Y/%m/%d')

intents = discord.Intents.all() #デフォルトのインテンツオブジェクトを生成
bot = commands.Bot(command_prefix='?', help_command = None, description=description, intents=intents)

#採掘用変数
MineResult_Show = []
Mine_Location = {"f1":"都市近郊（Lv0から利用可能）", "f2":"ミドガルネ北方（Lv1から利用可能）", "f3":"ミドガルネ南方（Lv2から利用可能）", "f4":"竜山地帯（Lv3から利用可能）", "f5":"ドラゴンズエッジ（Lv4から利用可能）", "f6":"古洞窟",
                 "n1":"ごみ捨て場（Lv0から利用可能）","n2":"廃品置き場（Lv1から利用可能）","n3":"棄てられた鉱山（Lv2から利用可能）","n4":"山中の坑道（Lv3から利用可能）","n5":"台地の採掘場（Lv4から利用可能）"}
Mine_result = {"f1":{"1":"石ころ", "2":"硝石", "3":"鉄鉱石", "4":"銅鉱石", "5":"金鉱石"},
               "f2":{"6":"石ころ", "7":"鉄鉱石", "8":"銅鉱石", "9":"銀鉱石", "10":"金鉱石"},
               "f3":{"11":"石ころ", "12":"鉄鉱石", "13":"銅鉱石", "14":"銀鉱石", "15":"月長石","16":"金鉱石"},  
               "f4":{"17":"石ころ", "18":"鉄鉱石", "19":"銅鉱石", "20":"銀鉱石", "21":"蒼鉄晶","22":"金鉱石"}, 
               "f5":{"23":"石ころ", "24":"鉄鉱石", "25":"銅鉱石", "26":"銀鉱石", "27":"蒼鉄晶", "28":"アルブム・クリスタル", "29":"金鉱石"},
               "f6":{"30":"石ころ", "31":"鉄鉱石", "32":"銅鉱石", "33":"銀鉱石", "34":"蒼鉄晶", "35":"灰銀石"},
               "n1":{"36":"石ころ", "37":"スクラップ"},
               "n2":{"38":"石ころ", "39":"スクラップ", "40":"アルミの破片", "41":"鉄の破片"},
               "n3":{"42":"石ころ", "43":"スクラップ", "44":"アルミの破片", "45":"鉄の破片", "46":"銅の破片"},
               "n4":{"47":"石ころ", "48":"アルミの破片", "49":"石炭", "50":"銅の破片", "51":"銀の破片"},
               "n5":{"52":"石ころ", "53":"アルミの破片", "54":"鉄の破片", "55":"銅の破片", "56":"銀の破片", "57":"チタンの破片"}
               }
MineLocData = {"f1":{"Start":1,"End":5}, "f2":{"Start":6, "End":10}, "f3":{"Start":11, "End":16}, "f4":{"Start":17, "End":22},"f5":{"Start":23, "End":29},"f6":{"Start":30, "End":35},
               "n1":{"Start":36,"End":37}, "n2":{"Start":38, "End":41}, "n3":{"Start":42, "End":46}, "n4":{"Start":47, "End":51},"n5":{"Start":52, "End":57}}
Mine_List = {"1":{"LL":1, "UL":65}, "2":{"LL":65, "UL":80}, "3":{"LL":81, "UL":95}, "4":{"LL":96, "UL":99}, "5":{"LL":100,"UL":100},
             "6":{"LL":1, "UL":70}, "7":{"LL":71, "UL":85}, "8":{"LL":86, "UL":95}, "9":{"LL":96, "UL":99}, "10":{"LL":100,"UL":100},
             "11":{"LL":1, "UL":60}, "12":{"LL":61, "UL":70}, "13":{"LL":71, "UL":80}, "14":{"LL":81, "UL":95}, "15":{"LL":96,"UL":99}, "16":{"LL":100,"UL":100},
             "17":{"LL":1, "UL":50}, "18":{"LL":51, "UL":70}, "19":{"LL":71, "UL":80}, "20":{"LL":81, "UL":95}, "21":{"LL":96,"UL":99}, "22":{"LL":100,"UL":100},
             "23":{"LL":1, "UL":40}, "24":{"LL":41, "UL":60}, "25":{"LL":61, "UL":70}, "26":{"LL":71, "UL":80}, "27":{"LL":81,"UL":95}, "28":{"LL":96,"UL":99}, "29":{"LL":100,"UL":100},
             "30":{"LL":1, "UL":40}, "31":{"LL":41, "UL":60}, "32":{"LL":61, "UL":70}, "33":{"LL":71, "UL":80}, "34":{"LL":81,"UL":95}, "35":{"LL":96,"UL":100},
             "36":{"LL":1, "UL":69}, "37":{"LL":70, "UL":100},
             "38":{"LL":1, "UL":70}, "39":{"LL":71, "UL":85}, "40":{"LL":86, "UL":95}, "41":{"LL":96, "UL":100},
             "42":{"LL":1, "UL":60}, "43":{"LL":61, "UL":70}, "44":{"LL":71, "UL":80}, "45":{"LL":81, "UL":95}, "46":{"LL":96, "UL":100},
             "47":{"LL":1, "UL":50}, "48":{"LL":51, "UL":70}, "49":{"LL":71, "UL":80}, "50":{"LL":81, "UL":95}, "51":{"LL":96, "UL":100},
             "52":{"LL":1, "UL":40}, "53":{"LL":41, "UL":50}, "54":{"LL":51, "UL":70}, "55":{"LL":71, "UL":80}, "56":{"LL":81, "UL":89}, "57":{"LL":90, "UL":100}
            }
#採取用変数
GatherResult_Show = []
Gather_Location = {"f1":"都市近郊の草原（Lv0から利用可能）", "f2":"エール湖湖畔（Lv1から利用可能）", "f3":"白亜の森（Lv2から利用可能）", "f4":"休火山の麓（Lv0から利用可能）", "f5":"妖精の森（Lv4から利用可能）",
                   "n1":"中部の草原（Lv0から利用可能）", "n2":"🌲北方の湖畔（Lv1から利用可能）","n3":"南方の砂浜（Lv2から利用可能）","n4":"西部の荒地（Lv3から利用可能）","n5":"ウェストキャニオン（Lv4から利用可能）"}
Gather_result = {"f1":{"1":"雑草", "2":"キノコ", "3":"薬草", "4":"タマゴ"},
                 "f2":{"5":"雑草", "6":"カワガ二", "7":"清水"},
                 "f3":{"8":"雑草", "9":"薬草", "10":"青癒草", "11":"カカオ"},
                 "f4":{"12":"石ころ", "13":"硫黄", "14":"魔水晶の破片"},
                 "f5":{"15":"石ころ", "16":"緑癒草", "17":"ヒラタケ", "18":"妖精トンボ", "19":"マボロシドングリ", "20":"マンドラゴラ"},
                 "n1":{"21":"雑草", "22":"キノコ", "23":"ニンジン", "24":"タマゴ"},
                 "n2":{"25":"雑草", "26":"キノコ", "27":"清水", "28":"タマゴ"},
                 "n3":{"29":"雑草", "30":"木材(流木)", "31":"カカオ"},
                 "n4":{"32":"雑草", "33":"サボテン", "34":"原油"},
                 "n5":{"35":"雑草", "36":"サボテン", "37":"原油"}}
GatherLocData = {"f1":{"Start":1, "End":4}, "f2":{"Start":5, "End":7}, "f3":{"Start":8, "End":11}, "f4":{"Start":12, "End":14}, "f5":{"Start":15,"End":20},
                 "n1":{"Start":21, "End":24}, "n2":{"Start":25, "End":28}, "n3":{"Start":29, "End":31}, "n4":{"Start":32, "End":34}, "n5":{"Start":35, "End":37}}
Gather_List = {"1":{"LL":1, "UL":50}, "2":{"LL":51, "UL":70}, "3":{"LL":71, "UL":90}, "4":{"LL":91, "UL":100},
               "5":{"LL":1,"UL":50},"6":{"LL":51, "UL":70}, "7":{"LL":71, "UL":100},
               "8":{"LL":1, "UL":50}, "9":{"LL":51, "UL":70}, "10":{"LL":71,"UL":90},"11":{"LL":91, "UL":100},
               "12":{"LL":1, "UL":80}, "13":{"LL":81, "UL":95}, "14":{"LL":96, "UL":100},
               "15":{"LL":1, "UL":50}, "16":{"LL":51, "UL":60}, "17":{"LL":61, "UL":70},"18":{"LL":71, "UL":80}, "19":{"LL":81, "UL":95}, "20":{"LL":96,"UL":100},
               "21":{"LL":1, "UL":50}, "22":{"LL":51, "UL":70}, "23":{"LL":71, "UL":90}, "24":{"LL":91, "UL":100},
               "25":{"LL":1, "UL":50}, "26":{"LL":51, "UL":70}, "27":{"LL":71, "UL":90}, "28":{"LL":91, "UL":100},
               "29":{"LL":1, "UL":50}, "30":{"LL":51, "UL":90}, "31":{"LL":91, "UL":100},
               "32":{"LL":1, "UL":80}, "33":{"LL":81, "UL":90}, "34":{"LL":91, "UL":100},
               "35":{"LL":1, "UL":70}, "36":{"LL":71, "UL":90}, "37":{"LL":91, "UL":100}
               }

#ゲーム制作関連
game_bug = {}
game_make = {1:{"SYS":0,"STR":0,"TEX":0,"ACT":0,"SND":0,"MLT":0}}
graphic = {"ドットゲームLv1":0,"ドットゲームLv2(✧)":0,"3DゲームLv1(✧✧)":0,"3DゲームLv2(✧✧✧)":0}
game_hard = {"ハウスコンピューター":{"SYS":0,"STR":0,"TEX":5,"ACT":0,"SND":5,"MLT":0},
             "アタレ1500":{"SYS":5,"STR":0,"TEX":0,"ACT":5,"SND":0,"MLT":0},
             "スーパーハウコン":{"SYS":0,"STR":0,"TEX":20,"ACT":0,"SND":20,"MLT":0},
             "ギガドライブ":{"SYS":10,"STR":0,"TEX":10,"ACT":10,"SND":0,"MLT":0},
             "サガリターン":{"SYS":20,"STR":0,"TEX":20,"ACT":15,"SND":15,"MLT":0},
             "PlayVerse":{"SYS":15,"STR":0,"TEX":15,"ACT":20,"SND":20,"MLT":0},
             "MANTENDO46":{"SYS":17,"STR":0,"TEX":17,"ACT":17,"SND":17,"MLT":0},
             "ドリームブロード":{"SYS":30,"STR":0,"TEX":30,"ACT":25,"SND":25,"MLT":0},
             "PlayVerse2":{"SYS":25,"STR":0,"TEX":25,"ACT":30,"SND":30,"MLT":0},
             "ゲームブロック":{"SYS":30,"STR":0,"TEX":25,"ACT":30,"SND":30,"MLT":0},
             }

game_junle = {"ブロック崩し/パズル":{"SYS":5,"STR":0,"TEX":5,"ACT":0,"SND":0,"MLT":0},
              "横スクロールアクション":{"SYS":5,"STR":0,"TEX":0,"ACT":5,"SND":0,"MLT":0},
              "縦スクロールシューティング":{"SYS":0,"STR":0,"TEX":0,"ACT":10,"SND":0,"MLT":0},
              "RPG":{"SYS":0,"STR":5,"TEX":0,"ACT":0,"SND":5,"MLT":0},
              "格闘ゲーム✧":{"SYS":10,"STR":0,"TEX":0,"ACT":10,"SND":0,"MLT":0},
              "SRPG✧":{"SYS":5,"STR":10,"TEX":0,"ACT":0,"SND":0,"MLT":0},
              "SFアクション✧":{"SYS":0,"STR":0,"TEX":10,"ACT":10,"SND":0,"MLT":0},
              "サウンドノベル✧":{"SYS":0,"STR":15,"TEX":0,"ACT":0,"SND":5,"MLT":0},
              "FRPG✧":{"SYS":0,"STR":10,"TEX":0,"ACT":0,"SND":10,"MLT":0},
              "シミュレーションゲーム✧":{"SYS":10,"STR":0,"TEX":10,"ACT":0,"SND":0,"MLT":0},
              "ミリタリーアクション✧✧":{"SYS":0,"STR":0,"TEX":15,"ACT":25,"SND":25,"MLT":0},
              "アクションRPG✧✧":{"SYS":0,"STR":20,"TEX":20,"ACT":20,"SND":0,"MLT":0},
              "ステルスアクション✧✧":{"SYS":15,"STR":15,"TEX":15,"ACT":15,"SND":0,"MLT":0},
              "黎明型3DFPS✧✧":{"SYS":25,"STR":0,"TEX":15,"ACT":25,"SND":0,"MLT":0},
              "MMORPG✧✧✧":{"SYS":0,"STR":20,"TEX":20,"ACT":20,"SND":0,"MLT":40},
              "初期型3DFPS✧✧✧":{"SYS":35,"STR":0,"TEX":25,"ACT":35,"SND":0,"MLT":0},
              "アクションRPG Ⅱ✧✧✧":{"SYS":0,"STR":30,"TEX":30,"ACT":25,"SND":25,"MLT":0},
              "SRPGⅡ✧✧":{"SYS":35,"STR":35,"TEX":0,"ACT":0,"SND":0,"MLT":0},
              "SRPGⅡ✧✧✧":{"SYS":55,"STR":55,"TEX":0,"ACT":0,"SND":0,"MLT":0},
              "ミリタリーアクションⅡ✧✧✧":{"SYS":0,"STR":5,"TEX":35,"ACT":35,"SND":35,"MLT":0},
              "FRPG Ⅱ✧✧✧":{"SYS":25,"STR":35,"TEX":25,"ACT":0,"SND":35,"MLT":0}}
game_trend = {"王道":{"SYS":25,"STR":0,"TEX":0,"ACT":0,"SND":0,"MLT":0},
              "ストーリー重視":{"SYS":0,"STR":25,"TEX":0,"ACT":0,"SND":0,"MLT":0},
              "動作重視":{"SYS":0,"STR":0,"TEX":0,"ACT":0,"SND":0,"MLT":0},
              "テクスチャ重視":{"SYS":0,"STR":0,"TEX":25,"ACT":0,"SND":0,"MLT":0},
              "アクション重視":{"SYS":0,"STR":0,"TEX":0,"ACT":25,"SND":0,"MLT":0},
              "バランス重視":{"SYS":5,"STR":5,"TEX":5,"ACT":5,"SND":5,"MLT":5}}

#農業関連
produce = {"トマト/野菜":{"needs":0.5,"harv":"1d10+5"},"キャベツ/野菜":{"needs":0.5,"harv":"1d10+5"},
           "白菜/野菜":{"needs":0.5,"harv":"1d10+5"},"ニンジン/野菜":{"needs":0.5,"harv":"1d10+5"},
           "ジャガイモ":{"needs":0.5,"harv":"1d10+5"},"たまねぎ/野菜":{"needs":0.5,"harv":"1d10+5"},
           "ネギ/野菜":{"needs":0.5,"harv":"1d10+5"},"カボチャ/野菜":{"needs":0.5,"harv":"1d6+2"},
           "麦/穀物":{"needs":0.5,"harv":"1d10+5"},"米/穀物":{"needs":0.5,"harv":"1d15+5"},
           "綿花/植物":{"needs":0.5,"harv":"1d6"},"アザミ/植物":{"needs":0.5,"harv":"1d3"},
           "れんげ/植物":{"needs":0.5,"harv":"1d3"},
           "獣肉(肉牛)/牧畜":{"needs":1,"harv":"1d10+5"},"ミルク(乳牛)/牧畜":{"needs":1,"harv":"1d10+5"},
           "鶏肉(茶鶏)/牧畜":{"needs":0.5,"harv":"1d8+2"},"タマゴ(白鶏)/牧畜":{"needs":0.5,"harv":"1d8+2"},
           "羊毛(白ヒツジ)/牧畜":{"needs":1,"harv":"1d3"},"生糸(カイコ)/牧畜":{"needs":0.5,"harv":"1d3"},}
square = {"0.5":0.5,"1":1,"2":2,"3":3,"4":4}
farm_user = {1:{"size":0.5,"veg":"トマト/野菜"}}

#戦闘管理コマンド
#---登録型---
current_quest = {"UID":{"QID":"f1",}}
current_enemy = {"UID":{"A":"Fn", "B":"Ded", "C":"Alv", "D":"Alv", "E":"Alv", "F":"Alv"}}
current_user = {}
#---固定型---
script_list = ["エル・プサイ・コングルゥ","ラ・ヨダソウ・スティアーナ","お前を見ているぞ","スーパーハカー","クリスティーナ","クリス","ツチノコ","チビ","ビッチ"]
enemy_tag = ["A","B","C","D","E","F","G","H","I"]
quest_id = {"f1":{"name":"スノーゴブリン討伐任務","loc":2,"col":1},
            "f2":{"name":"スノーウルフ討伐任務", "loc":3,"col":2},
            "f3":{"name":"碧爪は冷天を裂く", "loc":4, "col":3},
            "f4":{"name":"山明の地荒らせし幼竜", "loc":5, "col":4},
            "f5":{"name":"碧爪は冷天を裂く", "loc":6, "col":5},
            "f6":{"name":"ファングボア狩猟任務", "loc":7, "col":6}}

def calc_mine(result, loc):
    i = MineLocData[loc]["Start"]
    while i <= MineLocData[loc]["End"]: #採掘場所のStart値よりもEnd値の方が大きい場合は
        LL = Mine_List[str(i)]["LL"]
        UL = Mine_List[str(i)]["UL"]
        if LL <= result <= UL: #採掘場所のリザルトを全て試し、result値と該当するものがあれば結果に追加
            MineResult_Show.append((Mine_result[loc][str(i)]))
        else:
            pass
        i+=1

def calc_gather(result, loc):
    i = GatherLocData[loc]["Start"]
    while i <= GatherLocData[loc]["End"]: #採掘場所のStart値よりもEnd値の方が大きい場合は
        LL = Gather_List[str(i)]["LL"]
        UL = Gather_List[str(i)]["UL"]
        if LL <= result <= UL: #採掘場所のリザルトを全て試し、result値と該当するものがあれば結果に追加
            GatherResult_Show.append((Gather_result[loc][str(i)]))
        else:
            pass
        i+=1

class Questionnaire(ui.Modal, title='財産記録更新'):
    answer = ui.TextInput(label='内容', style=discord.TextStyle.paragraph)

    async def on_submit(self, interaction: discord.Interaction):
        usr_id = str(interaction.user.id)+"A"
        thread = interaction.guild.get_thread(th_name[usr_id])
        await interaction.response.send_message('財産記録を更新しました。',ephemeral=True)
        await update(usr_id,thread,text=self.answer)

class TrueButton(discord.ui.Button):
    def __init__(self, *, style: ButtonStyle = ButtonStyle.secondary, label:str = "更新",msg):
        super().__init__(style=style, label=label, disabled=False)
        self.label = label
        self.msg = msg

    async def callback(self, interaction: discord.Interaction):
        if self.label == "更新":
            modal = Questionnaire()
            await interaction.response.send_modal(modal)
            await interaction.message.delete()
        if self.label == "最新の記録を確認":
            await interaction.response.send_message(self.msg.content,ephemeral=True)

#ゲーム制作時、確定用ボタン
class GameButton(discord.ui.Button):
    def __init__(self, *, style: ButtonStyle = ButtonStyle.secondary, label:str = "確定",msg,name,ctx):
        super().__init__(style=style, label=label, disabled=False)
        self.name = name
        self.msg = msg
        self.ctx = ctx

    async def callback(self, interaction: discord.Interaction):
        await interaction.response.send_message("`>ゲーム開発を行います`")
        await self.msg.delete()
        await interaction.message.delete()
        await develop_game(self.ctx)

#ゲーム制作セレクトリスト
class GameList(discord.ui.Select):
    def __init__(self,listed,text,ctx):
        options=[]
        self.before = {"SYS":0,"STR":0,"TEX":0,"ACT":0,"SND":0,"MLT":0}
        self.dict = listed
        for sp_keys in listed:
            options.append(discord.SelectOption(label=sp_keys, description=''))
    
        super().__init__(placeholder=text, min_values=1, max_values=1, options=options)

    async def callback(self, interaction: discord.Interaction):
        uid = interaction.user.id
        if self.dict != graphic:
            game_make.update({uid:{"SYS":game_make[uid]["SYS"]-self.before["SYS"],
                                   "STR":game_make[uid]["STR"]-self.before["STR"],
                                   "TEX":game_make[uid]["TEX"]-self.before["TEX"],
                                   "ACT":game_make[uid]["ACT"]-self.before["ACT"],
                                   "SND":game_make[uid]["SND"]-self.before["SND"],
                                   "MLT":game_make[uid]["MLT"]-self.before["MLT"],}})
            #増加
            SYS = game_make[uid]["SYS"] + self.dict[self.values[0]]["SYS"]
            STR = game_make[uid]["STR"] + self.dict[self.values[0]]["STR"]
            TEX = game_make[uid]["TEX"] + self.dict[self.values[0]]["TEX"]
            ACT = game_make[uid]["ACT"] + self.dict[self.values[0]]["ACT"]
            SND = game_make[uid]["SND"] + self.dict[self.values[0]]["SND"]
            MLT = game_make[uid]["MLT"] + self.dict[self.values[0]]["MLT"]
            self.before = self.dict[self.values[0]]
            game_make.update({uid:{"SYS":SYS,"STR":STR,"TEX":TEX,"ACT":ACT,"SND":SND,"MLT":MLT}})
            print(game_make[uid])
        else:
            pass
        if self.values[0] == "動作重視":
            game_bug[uid] = 1
        else:
            game_bug[uid] = 0
        await interaction.response.defer()
        #await interaction.response.send_message("s",ephemeral=True)

#農業用セレクトメニュー
class Farming(discord.ui.Select):
    def __init__(self,listed,text,ctx):
        options=[]
        self.dict = listed
        for sp_keys in listed:
            options.append(discord.SelectOption(label=sp_keys, description=''))
        uid = ctx.author.id
        farm_user.update({uid:{"size":0,"veg":"default"}})
    
        super().__init__(placeholder=text, min_values=1, max_values=1, options=options)

    async def callback(self, interaction: discord.Interaction):
        uid = interaction.user.id
        if self.dict == square:
            farm_user[uid]["size"] = self.dict[self.values[0]] #例えば"0.5"が代入される
        elif self.dict == produce:
            farm_user[uid]["veg"] = self.values[0] #例えば"ニンジン"が代入される
        else:
            pass
        await interaction.response.defer()

#農業確定ボタン
class FarmButton(discord.ui.Button):
    def __init__(self, *, style: ButtonStyle = ButtonStyle.secondary, label:str = "農業開始",msg,ctx,text):
        super().__init__(style=style, label=label, disabled=False)
        self.msg = msg
        self.ctx = ctx
        self.text = text

    async def callback(self, interaction: discord.Interaction):
        await interaction.response.send_message("`24/48時間後に収穫結果が通知されます`")
        uid = interaction.user.id
        print(farm_user[uid])
        await self.msg.delete()
        await interaction.message.delete()
        await farm_result(self.ctx,self.text)

#チャットパレット登録用モーダル
class regpallet(ui.Modal, title='新規チャットパレット登録'):
    charaname = ui.TextInput(label='キャラ名', style=discord.TextStyle.short)
    com = ui.TextInput(label="コマンド", style=discord.TextStyle.short)

    async def on_submit(self, interaction: discord.Interaction):
        usr_id = str(interaction.user.id)
        str(self.charaname)
        print(type(self.com),self.charaname)
        COMdatasheet.append_row(['null',usr_id,str(self.charaname),str(self.com)])
        await interaction.response.send_message('`チャットパレットを登録しました`',ephemeral=True)

#コマンドリスト用クラス
class MakeList(discord.ui.View):
    def __init__(self,usr_id):
        super().__init__()
        self.add_item(CommandList(usr_id=usr_id))

class CommandList(discord.ui.Select): #セレクトメニューの選択肢を追加するクラス
    def __init__(self,usr_id):
        COMdf = pd.DataFrame(COMdatasheet.get_all_values()[1:],columns=COMdatasheet.get_all_values()[0])
        COMdfA = COMdf.set_index("ユーザーID")
        chatpaletts = COMdfA.loc[usr_id]
        options_list=[]
        for item in chatpaletts:
            options_list.append(discord.SelectOption(label=item, description=''))
    
        super().__init__(placeholder='', min_values=1, max_values=1, options=options_list)

    async def callback(self, interaction: discord.Interaction):
        await interaction.response.send_message(f"{self.values[0]}", ephemeral=True)

@bot.tree.command(name="npc_say",description="仮想ユーザーで発言します")
async def test_command(interaction: discord.Interaction,text:str):
    Udf = pd.DataFrame(Fdatasheet.get_all_values()[1:],columns=Fdatasheet.get_all_values()[0])
    usr_id = str(interaction.user.id)+"H"
    if usr_id not in str(Udf["UID"]):
        return
    else:
        UdfA = Udf.set_index("UID")
        channel = interaction.channel
        avatar = UdfA.at[str(usr_id), "avatar"]
        names = UdfA.at[str(usr_id), "name"]
        webhook = await channel.create_webhook(name="試験用")
        await webhook.send(content=text,username=names,wait=True,avatar_url=avatar)
        await interaction.response.send_message("`メッセージを発言しました`",ephemeral=True)
        await asyncio.sleep(10)
        #await whmsg.delete()
        await webhook.delete()
        print("Webhookを削除")

@bot.tree.command(name="クエスト開始",description="クエストを開始します。")
@app_commands.describe(qid="クエストIDを入力してください",front="前衛を入力してください(複数の場合は 、で区切り)",back="後衛を入力してください(複数の場合は 、で区切り)")
async def battle_start(interaction: discord.Interaction,qid:typing.Literal["f1","f2","f3","f4","f5","f6"], front:str = "none.", back:str = "none."):
    name="name"
    usr_id = str(interaction.user.id) #ユーザーIDを取得
    if usr_id in blacklist:
        #await interaction.response.send_message("discord.ext.commands.errors.CommandInvokeError: Command raised an exception: HTTPException: 400 Bad Request (error code: 50006): Cannot send message")
        return
    if usr_id in current_quest.keys():
        await interaction.response.send_message("> 既にクエストを開始しています")
        return 
    current_quest[usr_id] = {"qid":qid} #クエストIDを登録
    usr_num = 0
    if "、" in front: #もし前衛が複数人いればリスト化
        front_list = front.split("、")
        for i in range(len(front_list)):
            add_dict = current_quest[usr_id]
            add_dict[front_list[i]] = "front"
            current_quest[usr_id] = add_dict
            usr_num += 1
    else: #前衛が一人の場合
            add_dict = current_quest[usr_id]
            add_dict[front] = "front"
            current_quest[usr_id] = add_dict
            usr_num += 1
    if "、" in back: #もし後衛が複数人いればリスト化
        back_list = back.split("、")
        for b in range(len(back_list)):
            add_dict = current_quest[usr_id]
            add_dict[back_list[b]] = "back"
            current_quest[usr_id] = add_dict
            print(current_quest[usr_id])
            usr_num += 1
    elif "none" in back: #後衛がいなければスキップ
        pass
    else: #後衛が一人の場合
            add_dict = current_quest[usr_id]
            add_dict[back] = "back"
            current_quest[usr_id] = add_dict
            usr_num += 1
    #敵数追加処理
    wb = openpyxl.load_workbook("quests.xlsx")
    based = wb.worksheets[0]
    eny_base = based.cell(quest_id[qid]["loc"], 3).value
    eny_add = based.cell(quest_id[qid]["loc"], 4).value
    eny_num = int(eny_base) + int(eny_add) * (usr_num-1)
    current_enemy[usr_id] = {"eny":str(eny_num)}
    for c in range(eny_num):
         add_dict = current_enemy[usr_id]
         add_dict[enemy_tag[c]] = "Alv"
         current_enemy[usr_id] = add_dict
    print(current_quest[usr_id])
    print(current_enemy[usr_id])
    await interaction.response.send_message(f"`クエストID: {quest_id[qid][name]} |クエスト人数: {usr_num}`\n`前衛: {front}`\n`後衛: {back}`\n`進行担当者: {interaction.user.display_name}`",ephemeral=False)
    wb.close()

@bot.tree.command(name="チャットパレット",description="チャットパレットです")
@app_commands.describe(cnt="操作を選択してください",cmnd="登録するコマンドを入力してください")
async def chatpallet(interaction: discord.Interaction, cnt:typing.Literal["表示","登録"], cmnd:str = "default"):
    usr_id = str(interaction.user.id)
    if usr_id in COMdatasheet.col_values(2): #ユーザーIDが登録済みであれば
        if cnt == "登録": #コマンド追加の場合
            target = COMdatasheet.find(str(usr_id)) #登録したUIDからターゲット座標を取得
            T_ROW_ARRAY = COMdatasheet.row_values(target.row)
            LAST_ROW_IDX = len(T_ROW_ARRAY)
            print(T_ROW_ARRAY,LAST_ROW_IDX)
            COMdatasheet.update_cell(target.row,LAST_ROW_IDX+1,str(cmnd))
            del usr_id
            await interaction.response.send_message("`チャットパレットに登録しました`", ephemeral=True)
            return
        elif cnt == "表示":        
            await interaction.response.send_message(view=MakeList(usr_id), ephemeral=True)
            del usr_id
        else:
            pass
    else: #新規登録の場合
        modal = regpallet()
        await interaction.response.send_modal(modal)
        del usr_id

@bot.event
async def on_ready():
    print(f'ログインしました: {bot.user} (ID: {bot.user.id})')
    print(now)
    await bot.tree.sync()
    channel = bot.get_channel(1085309399955943454)
    await channel.send(f"起動時間：{dad}\nGatherV1.3を起動しました。\nV1.1")
    print('------')

#採掘コマンド
@bot.command()
async def mine(ctx, loc:str, num: int,skill: int):
    usr_id = str(ctx.author.id)
    if usr_id in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    suc = 0
    format = ""
    usr_nm = ctx.author.display_name
    for i in range(num): #num回数分繰り返し、
        Dresult = random.randrange(1,101,1)
        if Dresult <= skill:
            Mresult = random.randrange(1,101,1) #掘削結果
            calc_mine(Mresult,loc)
            suc += 1
        else:
            pass
    for item in MineResult_Show:
        format += (item + "、")
    await ctx.send("**採掘場所: " + Mine_Location[loc] + "**")
    await ctx.send("`" + format + f"|採掘終了`\n> 採掘者:{usr_nm} | 成功回数:" + str(suc))
    MineResult_Show.clear()

#採取コマンド
@bot.command()
async def gather(ctx, loc:str, num: int,skill: int):
    usr_id = str(ctx.author.id)
    if usr_id in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    suc = 0
    format = ""
    usr_nm = ctx.author.display_name
    for i in range(num): #num回数分繰り返し、
        Dresult = random.randrange(1,101,1) #技能結果
        if Dresult <= skill:
            Mresult = random.randrange(1,101,1) #採取結果
            calc_gather(Mresult,loc)
            suc += 1
        else:
            pass
    for item in GatherResult_Show:
        format += (item + "、")
    print(GatherResult_Show)
    await ctx.send("**採取場所: " + Gather_Location[loc] + "**")
    await ctx.send("`" + format + f"|採取終了`\n> 採取者:{usr_nm} | 成功回数:" + str(suc))
    GatherResult_Show.clear()

#財産記録管理コマンド
@bot.command()
async def asset(ctx, Threadname:str="default"):
    #初期定義
    usr_id = str(ctx.author.id)+"A"
    if str(ctx.author.id) in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    channel = bot.get_channel(941289343014797333)
    member_mention = f"<@{ctx.author.id}>"
    admin_mention = f"<@1024785159821729792>"
    userflame = pd.DataFrame(datasheet.get_all_values()[1:],columns=datasheet.get_all_values()[0])
    ufA = userflame.set_index("UID")

    if str(usr_id) in str(userflame["UID"]):
        thread_id = int(ufA.at[str(usr_id),"スレッドID"])
        print(thread_id)
        th_name[usr_id] = thread_id
        thread = ctx.guild.get_channel_or_thread((thread_id))
        msg_id = thread.last_message_id
        msg = await thread.fetch_message(msg_id)
        view = discord.ui.View(timeout=None)
        view.add_item(TrueButton(label="最新の記録を確認",msg=msg))
        view.add_item(TrueButton(msg=msg))
        await ctx.send(view=view)
    else:
        if Threadname == "default":
            await ctx.send("`スレッド名を入力してください`")
            return
        thread = await channel.create_thread(name=Threadname)
        await thread.send(f"{admin_mention}{member_mention}:財産記録用スレッドです。")
        datasheet.append_row(values=[str(usr_id),],table_range='L2')
        target = datasheet.find(str(usr_id))
        print(target)
        datasheet.update_cell(target.row,target.col+1,str(thread.id))

async def update(usr_id,thread,text):
    naiyou = str(text)
    await thread.send(naiyou)
    del th_name[usr_id]

#ゲーム制作コマンド
@bot.command()
async def game(ctx, skill:str):
    if str(ctx.author.id) in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    result = random.randrange(1,101,1)
    if result > int(skill):
        await ctx.send("`>開発失敗`")
        return
    else:
        pass
    usr_id = ctx.author.id
    game_make.update({usr_id:{"SYS":0,"STR":0,"TEX":0,"ACT":0,"SND":0,"MLT":0}})
    menu = discord.ui.View(timeout=None)
    menu.add_item(GameList(graphic,"グラフィック種別",ctx))
    menu.add_item(GameList(game_hard,"ハード",ctx))
    menu.add_item(GameList(game_junle,"ジャンル",ctx))
    menu.add_item(GameList(game_trend,"方向性",ctx))
    msg = await ctx.send(view=menu)
    button = discord.ui.View(timeout=None)
    button.add_item(GameButton(msg=msg,name=(ctx.author.display_name),ctx=ctx))
    await ctx.send(view=button)

#制作用関数
async def develop_game(ctx):
    lot = 1 * 5000
    usr_id = ctx.author.id
    SYS = random.randrange(1,101,1) + game_make[usr_id]["SYS"]
    STR = random.randrange(1,101,1) + game_make[usr_id]["STR"]
    TEX = random.randrange(1,101,1) + game_make[usr_id]["TEX"]
    ACT = random.randrange(1,101,1) + game_make[usr_id]["ACT"]
    SND = random.randrange(1,101,1) + game_make[usr_id]["SND"]
    MLT = game_make[usr_id]["MLT"]
    bug = random.randrange(1,401,1)
    if game_bug[usr_id] == "1":
        bug = bug//2
    rieki = (SYS+STR+TEX+ACT+SND+MLT)*10 - (bug*10)
    lot = (SYS+STR+TEX+ACT+SND+MLT)*251
    res = f"システム: {SYS}\nストーリー: {STR}\nテクスチャ: {TEX}\nアクション: {ACT}\nサウンド: {SND}\nマルチ: {MLT}"

    embed_title=discord.Embed(title="ゲーム開発結果",
                        description=f"**開発者:**{ctx.author.display_name}```{res}```",
                        color=0x6E6636) 
    embed_title.add_field(name="`販売本数:`", value=lot, inline=True)
    embed_title.add_field(name="`利益(制作費抜き):`", value=rieki, inline=True)

    await ctx.send(embed=embed_title)
    game_bug.clear()
    game_make.clear()

#農業用コマンド
@bot.command()
async def farm(ctx, text:str):
    if str(ctx.author.id) in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    farm_menu = discord.ui.View(timeout=None)
    farm_menu.add_item(Farming(square,"マスを選択",ctx))
    farm_menu.add_item(Farming(produce,"作物を選択",ctx))
    msg = await ctx.send(view=farm_menu)
    button = discord.ui.View(timeout=None)
    button.add_item(FarmButton(msg=msg,ctx=ctx,text=text))
    await ctx.send(view=button)

#農業結果用関数
async def farm_result(ctx,text):
    grows = 0
    usr_id = ctx.author.id
    whats = farm_user[usr_id]["veg"]
    howlarge = farm_user[usr_id]["size"]
    number = produce[whats]["needs"]
    num_times = howlarge / number
    if "牧畜" in whats:
        await asyncio.sleep(172800)
    else:
        await asyncio.sleep(86400)
    member_mention = f"<@{usr_id}>"
    rolls = produce[whats]["harv"]
    for i in range(int(num_times)):
        grows += dice.roll(rolls) #エラー出てるけど問題なし
        print(grows)
    await ctx.send(f"{member_mention}\n`{whats}を{grows}個収穫しました。`\n`説明:{text}`\n`使用マス:{howlarge}`")
    del farm_user[usr_id]

#ブラックリストコマンド
@bot.command()
async def black(ctx, mode:str = "表示", ID:str=""):
    global blacklist
    if mode == "追加":
        blacklist.append(ID)
    elif mode == "削除":
        blacklist.remove(ID)
    elif mode == "表示":
        intel = ""
        for i in range(len(blacklist)):
            usr_id = int(blacklist[i])
            user = await bot.fetch_user(usr_id)
            intel += f"gather - {i}. ユーザー名:{user.name}{user.discriminator}\n "
        await ctx.send(intel)

#ガチャコマンド
@bot.command()
async def gacha(ctx, num:int, mode:str = "2", text:str = "なし"):
    global target
    usr_id = str(ctx.author.id)
    if str(ctx.author.id) in blacklist:
        return
    member_mention = f"<@{ctx.author.id}>"
    result = 0
    dice_list = []
    list_suc = [6,7,8,9,10,11,12,13,14,15,16,17,18,19,20]
    list_5 = [1,2,3,4,5]
    result_list = []
    modes = {"1":"1d100","2":"1d200","3":"1d250","4":"デバッグ"}
    if mode == "2":
        for i in range(num):
            result = dice.roll("1d300+0")
            dice_list.append(result)
    elif mode == "1":
        for i in range(num):
            result = dice.roll("1d200+0")
            dice_list.append(result)
    elif mode == "3":
        for i in range(num):
            result = dice.roll("1d250+0")
            dice_list.append(result)
    elif mode == "4":
        await ctx.send("デバッグモードです")
        dice_list = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20]
    elif mode != "1" and mode != "2" and mode != "3" and mode != "4":
        await ctx.send("不正なモードです")
        return
    if usr_id in blacklist:
        dice_list = [100,100]
    if usr_id == target:
        dice_list.append(1)
        dice_list.append(1)
        dice_list.append(2)
        dice_list.append(5)
    cur_mode = modes[mode]
    await ctx.send(f"> {num}回ガチャを振ります\n`ガチャ使用者:`{member_mention}\n`モード: {cur_mode} | 備考: {text}`")
    dept = (set(dice_list) & set(list_suc))
    dept_5 = (set(dice_list) & set(list_5))
    if dept != set():
        result_list = list(dept)
        await ctx.send(f"`6～20の結果:`\n"+', '.join(map(str,result_list)))
    if dept_5 != set():
        result_list = list(dept_5)
        await ctx.send(f"`5以下の結果:`\n"+', '.join(map(str,result_list)))
    if dept == set() and dept_5 == set():
        await ctx.send("`20以下の結果: なし`")
    else:
        pass

@bot.command()
async def gachatarget(ctx,target_ID:str):
    global target
    target = target_ID
    await ctx.send("ターゲットセット")

#NPCコマンド
@bot.command()
async def npc(ctx,text:str = "`発言内容が入力されていません`", name:str = "default"):
    usr_ids = str(ctx.author.id)
    if usr_ids in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    usr_id = str(ctx.author.id)+"H"
    Udf = pd.DataFrame(Fdatasheet.get_all_values()[1:],columns=Fdatasheet.get_all_values()[0])
    mesg = ctx.message
    if usr_id not in str(Udf["UID"]):
        channel = bot.get_channel(1081260969256296528)
        icon = ctx.message.attachments #添付ファイルを取得
        fp = await icon[0].to_file(filename=name+".png") #添付ファイルを送信可能な画像ファイルに変換
        msg = await channel.send(file=fp) #添付ファイルを指定チャンネルに投稿
        reg_url = msg.attachments[0].url #投稿したアバター画像のurlを取得
        Fdatasheet.append_row(values=[str(usr_id),],table_range='K2') #UIDをK2列に入力
        target = Fdatasheet.find(str(usr_id)) #登録したUIDからターゲット座標を取得
        Fdatasheet.update_cell(target.row,target.col+1,str(reg_url))
        Fdatasheet.update_cell(target.row,target.col+2,str(name))
        await ctx.send("> 仮想キャラを登録しました。")
    else:
        await mesg.delete()
        UdfA = Udf.set_index("UID")
        avatar = UdfA.at[str(usr_id), "avatar"]
        names = UdfA.at[str(usr_id), "name"]
        await sender(ctx,text,names,avatar)

#NPCコマンド用コマンド
@commands.command()
async def sender(ctx,text:str ,names:str, avatar:str):
    webhook = await ctx.channel.create_webhook(name="試験用")
    whmsg = await webhook.send(content=text,username=names,wait=True,avatar_url=avatar)
    await asyncio.sleep(10)
    #await whmsg.delete()
    await webhook.delete()
    print("Webhookを削除")

#満腹度管理コマンド
@bot.command()
async def food(ctx, world:str, plus:str = "0"):
    usr_id = str(ctx.author.id)
    if usr_id in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    #もし増加量が3以上であればリターン
    if int(plus) >= 4: 
        await ctx.send("`3以上は入力できません`")
        return
    #ファンタジアの場合
    if world == "f":
        Fusr = pd.DataFrame(Fdatasheet.get_all_values()[1:],columns=Fdatasheet.get_all_values()[0])
        FuserA = Fusr.set_index("ユーザーID")
        hung = FuserA.at[usr_id,"満腹度"]
        #増加量0の場合は現在値を表示
        if plus == "0":
            await ctx.reply(f"`ファンタジア | 現在の満腹度:{hung}/3.0`")
        #増加量がある場合は増加
        else:
            final = round((float(hung) + float(plus)),1)
            print(final)
            if final > 3:
                await ctx.send("`上限を超えています`")
                return
            target_ind = (FuserA.index.get_loc(usr_id)) + 2
            target_clm = 10
            Fdatasheet.update_cell(target_ind,target_clm,str(final))
            await ctx.reply(f"`ファンタジアの満腹度を{plus}回復させました。| 現在の満腹度:{final}/3.0`")
            await food_minus(usr_id,world)
    #ノクターンの場合
    elif world == "n":
        Nusr = pd.DataFrame(datasheet.get_all_values()[1:],columns=datasheet.get_all_values()[0])
        NuserA = Nusr.set_index("ユーザーID")
        hung = NuserA.at[usr_id,"満腹度"]
        #増加量0の場合は現在値を表示
        if plus == "0":
            await ctx.reply(f"`ノクターン | 現在の満腹度:{hung}/3.0`")
        #増加量がある場合は増加
        else:
            final = round((float(hung) + float(plus)),1)
            print(final)
            if final > 3:
                await ctx.send("`上限を超えています`")
                return
            target_ind = (NuserA.index.get_loc(usr_id)) + 2
            target_clm = 10
            datasheet.update_cell(target_ind,target_clm,str(final))
            await ctx.reply(f"`ノクターンの満腹度を{plus}回復させました。| 現在の満腹度:{final}/3.0`")
            await food_minus(usr_id,world)
    else:
        await ctx.send("`コマンドが不正です`")

async def food_minus(usr_id,world):
    if usr_id in food_count:
        return
    else:
        food_count.append(usr_id) #リストに追加
        while True:
            await asyncio.sleep(14400) #4時間待つ
            #ノクターンの場合
            if world == "n":
                Nusr = pd.DataFrame(datasheet.get_all_values()[1:],columns=datasheet.get_all_values()[0])
                NuserA = Nusr.set_index("ユーザーID")
                hung = NuserA.at[usr_id,"満腹度"]
                current_hung = round((float(hung) - float(0.5)),1)
                if current_hung < 0: #もし減少後の値が0以下なら
                    food_count.remove(usr_id)
                    break
                target_ind = (NuserA.index.get_loc(usr_id)) + 2
                target_clm = 10
                datasheet.update_cell(target_ind,target_clm,str(current_hung))
            #ファンタジアの場合
            elif world == "f":
                Fusr = pd.DataFrame(Fdatasheet.get_all_values()[1:],columns=Fdatasheet.get_all_values()[0])
                FuserA = Fusr.set_index("ユーザーID")
                hung = FuserA.at[usr_id,"満腹度"]
                current_hung = round((float(hung) - float(0.5)),1)
                if current_hung < 0: #もし減少後の値が0以下なら
                    food_count.remove(usr_id)
                    break
                target_ind = (FuserA.index.get_loc(usr_id)) + 2
                target_clm = 10
                Fdatasheet.update_cell(target_ind,target_clm,str(current_hung))
            #それ以外なら
            else:
                break

#bumpコマンド
@bot.command()
async def bump(ctx):
    target_channel = "939119300680228874"
    usr_id = ctx.author.id
    channel = str(ctx.channel.id)
    if str(usr_id) in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    print(channel)
    if target_channel == channel:       
        await ctx.send("2時間後に通知します")
        member_mention = f"<@{usr_id}>"
        await asyncio.sleep(7260) #2時間待つ
        await ctx.send(f"{member_mention}bump可能です")
    else:
        await ctx.send("bumpチャンネルではありません")

#戦闘管理コマンド
@bot.command()
async def ac(ctx, com:str, skill:str="0", num:str = "0"):
    print("a")
    usr_id = str(ctx.author.id)
    if usr_id in blacklist:
        #await ctx.send("discord.ext.error:The user is Persona non grata by faithlessness")
        return
    if usr_id in current_quest.keys():
        print("ユーザーIDあり")
        qid = current_quest[usr_id]["qid"]
        wb = openpyxl.load_workbook("quests.xlsx")
        if "回避" in com:
            target = com[0]
            evac_skill = skill
            for e in range(int(num)):
                evac = dice.roll("1d100+0")
                if evac > int(evac_skill): #失敗の場合
                    await ctx.send(f"> `{target}-{e}回目: 失敗`") 
                elif evac <= int(evac_skill): #成功の場合
                    await ctx.send(f"> `{target}-{e}回目: 成功`")
        elif "ターン" in com:
            keys = [k for k, v in current_enemy[usr_id].items() if v == 'Alv']
            fronts = [k for k, v in current_quest[usr_id].items() if v == 'front']
            based = wb.worksheets[0]
            detail = wb.worksheets[1]
            values = based.cell(quest_id[qid]["loc"], 5).value #行動選択肢の数を取得
            skill = based.cell(quest_id[qid]["loc"], 7).value #攻撃技能値を取得
            for i in range(len(keys)):
                atk = dice.roll("1d100+0")
                if 96 > atk > int(skill): #攻撃失敗の場合
                    await ctx.send(f"> `{keys[i]}: 失敗`") 
                    continue
                elif atk < int(skill)-5: #攻撃成功の場合
                    pass
                elif atk >= 96: #ファンブルの場合
                    self_dmg = dice.roll("1d10+0")
                    await ctx.send(f"> **{keys[i]}: ファンブル({self_dmg}の自傷ダメージ)**") 
                    continue
                elif int(skill)-5 <= atk <= int(skill): #クリティカル
                    await ctx.send(f"> **{keys[i]}: クリティカル(与ダメージ2倍)**") 
                act_selc = random.randrange(1,int(values)+1,1) #どの行動を選ぶかを処理
                act_detail = detail.cell(act_selc+1,quest_id[qid]["col"]).value
                action = act_detail.split("$")
                if "非攻撃" in action[0]: #攻撃以外なら
                    await ctx.send(f"> `{keys[i]}: {action[1]}`") 
                    return
                damage = dice.roll(action[0])
                if "全体" in action[1]: #全体攻撃であれば
                    await ctx.send(f"> `{keys[i]}: 全体に{damage}ダメージの攻撃` | `{action[1]}`") 
                elif len(fronts) != 1:
                    attack_target = random.randrange(1,len(fronts)+1,1)
                    if attack_target == 1:
                        attack_target = 0
                    else:
                        attack_target = 1
                    await ctx.send(f"> `{keys[i]}: {fronts[attack_target]}に{damage}ダメージの攻撃` | `{action[1]}`")
                    print("前衛二人")
                else:
                    await ctx.send(f"> `{keys[i]}: {fronts[0]}に{damage}ダメージの攻撃` | `{action[1]}`")
                    print("前衛一人")
                await asyncio.sleep(1)
            await ctx.send("`>ターン終了`")
        elif "終了" in com:
            del current_enemy[usr_id]
            del current_quest[usr_id]
            await ctx.send("> クエストを終了します")
        else:
             await ctx.send("> 不正なコマンドです")
        wb.close()
    else:
         await ctx.send("> クエストを開始していません")

#モンスター減少用コマンド
@bot.command()
async def dead(ctx, ID:str):
    target = ID.upper() 
    usr_id = str(ctx.author.id)
    if usr_id not in current_quest.keys():
        await ctx.send("> クエストを開始していません")
        return
    keys = [k for k, v in current_enemy[usr_id].items() if v == 'Alv']
    if target not in keys:
        await ctx.send("> 既に倒しています")
        return
    current_enemy[usr_id][target] = "ded"
    await ctx.send(f"> {ID}を倒しました")


#小ネタ
@bot.listen()
async def on_message(message):
    msg = message.content
    usr_id = str(message.author.id)
    dev = "1024785159821729792"
    #ブラックリスト対象はリターン
    if message.author == bot.user:
        return
    if usr_id in blacklist:
        return
    if msg not in script_list:
        return
    print("メッセージ")
    channel = message.channel
    webhook = await channel.create_webhook(name="試験用")
    if msg == "エル・プサイ・コングルゥ":
        avatar = "https://media.discordapp.net/attachments/1081260969256296528/1092341728851284009/gd_424703_-37.png"
        await webhook.send(content="エル・プサイ・コングルゥ",username="鳳凰院凶真",wait=True,avatar_url=avatar)
    elif msg == "ラ・ヨダソウ・スティアーナ":
        avatar = "https://media.discordapp.net/attachments/1081260969256296528/1092341728851284009/gd_424703_-37.png"
        await webhook.send(content="それがブッシュの選択...ってそれは元ネタの方ではないかっ！",username="鳳凰院凶真",wait=True,avatar_url=avatar)
    elif msg == "お前を見ているぞ":
        file = "https://media.discordapp.net/attachments/1081260969256296528/1092343258472009758/CRbZF2pU8AA5SIu.png"
        await channel.send(file)
    elif msg == "スーパーハカー":
        avatar = "https://media.discordapp.net/attachments/1081260969256296528/1092344604596121711/maxresdefault.jpg"
        await webhook.send(content="ハカーじゃなくてハッカーだろJK",username="ダル",wait=True,avatar_url=avatar)
    elif msg == "クリスティーナ":
        await channel.send("クリスティーナって呼ぶな！")
    elif msg == "クリス" and usr_id == dev:
        await channel.send("何か用？")
    elif msg == "ツチノコ":
        await webhook.send(content="スネーク、ツチノコを捕まえたのね！",username="パラメディック",wait=True,avatar_url="https://media.discordapp.net/attachments/1081260969256296528/1092350353481138177/FCdpEZcVcAEvsaF.jpg")
        await asyncio.sleep(2)
        await webhook.send(content="何だって！？",username="少佐",wait=True,avatar_url="https://media.discordapp.net/attachments/1081260969256296528/1092350353732808734/04.jpg")
        await asyncio.sleep(1)
        await webhook.send(content="ホントか、スネーク！？",username="シギント",wait=True,avatar_url="https://media.discordapp.net/attachments/1081260969256296528/1092350353950908416/EV7t5YxUwAMUKmt.jpg")
        await asyncio.sleep(1)
        await webhook.send(content="よくやった！さすがはザ・ボスの弟子だ！",username="シギント",wait=True,avatar_url="https://media.discordapp.net/attachments/1081260969256296528/1092350353950908416/EV7t5YxUwAMUKmt.jpg")
        await asyncio.sleep(2)
        await webhook.send(content="ああ、君を送り込んだ甲斐があったというものだ！",username="少佐",wait=True,avatar_url="https://media.discordapp.net/attachments/1081260969256296528/1092350353732808734/04.jpg")
        await asyncio.sleep(2)
        await webhook.send(content="さっさと任務を終わらせてそいつを連れ帰ってきてくれ。\n絶対食べたりするんじゃないぞ。いいな！",username="少佐",wait=True,avatar_url="https://media.discordapp.net/attachments/1081260969256296528/1092350353732808734/04.jpg")
        await asyncio.sleep(1)
    elif msg == "チビ":
        avatar ="https://media.discordapp.net/attachments/1081260969256296528/1092372313309196298/7A699337-DD8C-4FCA-A506-D98273375EC1.jpg"
        hook = await webhook.send(content="チビ言うな！",username="？？？",wait=True,avatar_url=avatar)
        await asyncio.sleep(2)
        await hook.delete()
    elif msg == "ビッチ":
        avatar ="https://media.discordapp.net/attachments/1081260969256296528/1092372313309196298/7A699337-DD8C-4FCA-A506-D98273375EC1.jpg"
        hook = await webhook.send(content="ビッチじゃない！",username="？？？",wait=True,avatar_url=avatar)
        await asyncio.sleep(2)
        await hook.delete()
    else:
        pass
    await webhook.delete()

bot.run('MTA3MjEwODcxNDAxNTg3MTAzNw.G6Tqbr.OQkLB7Gsm7VOohETEzBKGNq7OdmRDfaoTVfJmY')